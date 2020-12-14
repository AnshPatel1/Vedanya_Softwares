from calendar import monthrange
from datetime import datetime, date
from os import remove
from os.path import exists

import mysql.connector
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

rowCursorValue = 1

vhc_db = mysql.connector.connect(
    host="185.201.11.44",
    user="u257284371_ansh",
    password="ansh1Rutu"
)
cursor = vhc_db.cursor()


def get_all_transactions_by_client_on_month_year(client, month_year, start, end):
    print('hello')
    month_year = parser.parse('01' + month_year)
    month_year = month_year.strftime("%Y-%m-%d")
    month_year = datetime.strptime(month_year, "%Y-%m-%d")
    start_date = month_year.strftime("%Y-%m-%d")
    month = month_year.month - 1 + 1
    year = month_year.year + month // 12
    month = month % 12 + 1
    day = min(month_year.day, monthrange(year, month)[1])
    end_date = date(year, month, day).strftime("%Y-%m-%d")

    if datetime.strptime(start, "%Y-%m-%d") > datetime.strptime(start_date, "%Y-%m-%d"):
        start_date = start
    if datetime.strptime(end, "%Y-%m-%d") < datetime.strptime(end_date, "%Y-%m-%d"):
        end_date = end
    cursor.execute(
        "SELECT AMOUNT FROM u257284371_vedanya.transactions WHERE CLIENT_ID = {} AND DATE >= '{}' AND DATE < '{}';"
        "".format(get_client_id_by_name(client), start_date, end_date))
    list = cursor.fetchall()
    result = []
    if list:
        for item in list:
            result.append(item[0])
    return result


def make_convenient_client_by_month_transactions(months, client, start_date, end_date):
    result = {}
    max_row = 0
    print(months)
    for month in months:
        result[month] = get_all_transactions_by_client_on_month_year(client, month, start_date, end_date)
        if max_row < len(result[month]):
            max_row = len(result[month])
    return [max_row, result]


def get_client_id_by_name(name):
    cursor.execute("SELECT ID FROM u257284371_vedanya.clients WHERE NAME = '{}';".format(name))
    id = cursor.fetchone()
    return id[0]


def get_sr_id_by_name(name):
    cursor.execute("SELECT ID FROM u257284371_vedanya.sr WHERE NAME = '{}';".format(name))
    id = cursor.fetchone()
    return id[0]


def fetch_available_clients_by_sr_name(name):
    cursor.execute("SELECT NAME FROM u257284371_vedanya.clients WHERE SR_ID = {}".format(get_sr_id_by_name(name)))
    list = cursor.fetchall()
    result = []
    for item in list:
        result.append(item[0])
    return result


def monthList(dates):
    start, end = [datetime.strptime(_, "%Y-%m-%d") for _ in dates]
    total_months = lambda dt: dt.month + 12 * dt.year
    result = []
    for tot_m in range(total_months(start) - 1, total_months(end)):
        y, m = divmod(tot_m, 12)
        result.append(datetime(y, m + 1, 1).strftime("%b-%y"))
    return result


def rowCursorUpdate():
    global rowCursorValue
    rowCursorValue = rowCursorValue + 1


def split_years(start_date, end_date):
    date_list = []
    start_year = int(start_date[:4])
    end_year = int(end_date[:4])

    for year in range(start_year, end_year + 1):
        start = str(year) + '-01-01'
        end = str(year) + '-12-31'
        if datetime.strptime(start, "%Y-%m-%d") < datetime.strptime(start_date, "%Y-%m-%d"):
            start = start_date
        if datetime.strptime(end, "%Y-%m-%d") > datetime.strptime(end_date, "%Y-%m-%d"):
            end = end_date
        date_list.append([year, [start, end]])
    print(date_list)
    return date_list


def has_handle(fpath):
    try:
        remove(fpath)
        return False
    except PermissionError:
        return True


def generateReport(start_date, end_date, selection, save_path):
    print(selection)
    start_date = start_date.strftime("%Y-%m-%d")
    print(type(end_date))
    print(end_date)
    end_date = end_date.strftime("%Y-%m-%d")
    years = split_years(start_date, end_date)
    report = Workbook()
    report_sheet = report.active
    report_sheet.title = '2020'
    sheets = {}
    isFirstLoopCompleted = False
    for year in years:
        if not isFirstLoopCompleted:
            first_sheet = report.active
            first_sheet.title = str(year[0])
            sheets[year[0]] = first_sheet
            isFirstLoopCompleted = True
            continue
        sheet = report.create_sheet(str(year[0]))
        sheets[year[0]] = sheet

    for year_range in years:
        generateReportByDate(year_range[1][0], year_range[1][1], selection, sheets[year_range[0]])
    print(save_path)
    if exists(save_path):
        if has_handle(save_path):
            return 1
    report.save(save_path)
    return 0


def generateReportByDate(start_date, end_date, selection, report_sheet):
    staged_merge_cells = {}
    state_color = PatternFill(fill_type="solid", fgColor="FFFAAD")
    state_ttl_color = PatternFill(fill_type="solid", fgColor="FFFCC7")
    sr_color = PatternFill(fill_type="solid", fgColor="ADFFD1")
    sr_ttl_color = PatternFill(fill_type="solid", fgColor="CCFFE2")
    in_color = PatternFill(fill_type="solid", fgColor="A1FDFF")
    in_ttl_color = PatternFill(fill_type="solid", fgColor="D4FEFF")
    global rowCursorValue
    rowCursorValue = 1
    state_total_addresses = []
    state_month_addresses = {}
    month_cols = monthList([start_date, end_date])
    total_col_range = len(month_cols)

    report_sheet.merge_cells(start_row=rowCursorValue, start_column=1, end_row=rowCursorValue,
                             end_column=total_col_range + 4)
    _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 5, value='INDIA TOTAL')
    report_sheet['A' + str(rowCursorValue)].value = 'ALL INDIA'
    report_sheet['A' + str(rowCursorValue)].alignment = Alignment(horizontal='center')
    rowCursorUpdate()

    report_sheet['A' + str(rowCursorValue)].value = 'STATE'
    report_sheet['A' + str(rowCursorValue)].fill = state_color
    report_sheet.merge_cells(start_row=rowCursorValue, start_column=1, end_row=rowCursorValue,
                             end_column=total_col_range + 3)
    _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 4, value='STATE TOTAL')
    report_sheet[str(get_column_letter(total_col_range + 4)) + str(rowCursorValue)].fill = state_ttl_color
    rowCursorUpdate()
    rowCursorUpdate()
    if selection != {}:
        for state in selection.keys():
            if not selection[state]:
                report_sheet['A' + str(rowCursorValue)].value = str(state)
                report_sheet['A' + str(rowCursorValue)].fill = state_color
                report_sheet.merge_cells(start_row=rowCursorValue, start_column=1, end_row=rowCursorValue,
                                         end_column=total_col_range + 3)
                report_sheet[str(get_column_letter(total_col_range + 4)) + str(rowCursorValue)].fill = state_ttl_color
                rowCursorUpdate()
                continue

            report_sheet['A' + str(rowCursorValue)].value = str(state)
            report_sheet['A' + str(rowCursorValue)].fill = state_color
            report_sheet.merge_cells(start_row=rowCursorValue, start_column=1, end_row=rowCursorValue,
                                     end_column=total_col_range + 3)
            report_sheet[str(get_column_letter(total_col_range + 4)) + str(rowCursorValue)].fill = state_ttl_color
            rowCursorUpdate()
            sr_totals = []
            month_totals = {}
            for sr in selection[state]:

                sr_row = rowCursorValue
                double_reduction = False
                report_sheet['A' + str(rowCursorValue)].value = str(sr)
                report_sheet['A' + str(rowCursorValue)].fill = sr_color
                report_sheet.merge_cells(start_row=rowCursorValue, start_column=1, end_row=rowCursorValue,
                                         end_column=total_col_range + 2)
                _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 3, value='SR TOTAL')
                report_sheet[str(get_column_letter(total_col_range + 3)) + str(rowCursorValue)].fill = sr_ttl_color
                rowCursorUpdate()
                _ = report_sheet.cell(row=rowCursorValue, column=1, value='Clients')

                for items in month_cols:
                    _ = report_sheet.cell(row=rowCursorValue, column=month_cols.index(items) + 2, value=items)
                _ = report_sheet.cell(row=rowCursorValue, column=len(month_cols) + 2, value='TOTAL')
                rowCursorUpdate()
                start_ttl_vertical = rowCursorValue
                clients = fetch_available_clients_by_sr_name(sr)
                if clients:
                    for client in clients:
                        border_start = rowCursorValue
                        endx = 0
                        initial_pointer = rowCursorValue
                        start = rowCursorValue
                        data_package = make_convenient_client_by_month_transactions(month_cols, client, start_date,
                                                                                    end_date)
                        end = start + data_package[0] - 1
                        _ = report_sheet.cell(row=rowCursorValue, column=1, value=client)
                        if data_package[0] > 1:
                            staged_merge_cells['A' + str(rowCursorValue)] = [start, 1, end, 1]
                        for month in data_package[1].keys():
                            transactions = data_package[1][month]
                            for transaction in transactions:
                                _ = report_sheet.cell(row=initial_pointer, column=month_cols.index(month) + 2,
                                                      value=transaction)
                                if transactions.index(transaction) < len(transactions) - 1:
                                    initial_pointer = initial_pointer + 1
                            if endx <= initial_pointer:
                                endx = initial_pointer - 1
                            initial_pointer = rowCursorValue
                        _ = report_sheet.cell(row=initial_pointer, column=len(month_cols) + 2,
                                              value='=SUM({}:{})'.format(
                                                  str('B' + str(rowCursorValue)),
                                                  str(get_column_letter(len(month_cols) + 1) + str(endx))))

                        if not data_package[0] == 0:
                            double_reduction = True
                        if data_package[0] > 1:
                            rowCursorValue = rowCursorValue - 1
                        for _ in range(data_package[0] + 1):
                            rowCursorUpdate()

                rowCursorUpdate()
                _ = report_sheet.cell(row=rowCursorValue, column=1, value='TOTAL')
                for i in range(1, total_col_range + 4):
                    report_sheet[get_column_letter(i) + str(rowCursorValue)].fill = sr_ttl_color
                for i in range(sr_row, rowCursorValue + 1):
                    report_sheet[get_column_letter(total_col_range + 3) + str(i)].fill = sr_ttl_color

                for month in month_cols:
                    _ = report_sheet.cell(row=rowCursorValue, column=month_cols.index(month) + 2,
                                          value='=SUM({}:{})'.format(
                                              str(get_column_letter(month_cols.index(month) + 2)) + str(
                                                  start_ttl_vertical),
                                              str(get_column_letter(month_cols.index(month) + 2)) + str(
                                                  rowCursorValue - 1)))
                    if month not in month_totals.keys():
                        month_totals[month] = []

                    month_totals[month].append(
                        '{}{}'.format(str(get_column_letter(month_cols.index(month) + 2)), str(rowCursorValue)))

                _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 2,
                                      value='=SUM({}:{})'.format(
                                          'B' + str(rowCursorValue),
                                          str(get_column_letter(total_col_range + 1)) + str(rowCursorValue)))

                sr_totals.append(str(get_column_letter(total_col_range + 2)) + str(rowCursorValue))
                report_sheet.merge_cells(start_row=rowCursorValue, start_column=total_col_range + 2,
                                         end_row=rowCursorValue, end_column=total_col_range + 3)

                rowCursorUpdate()
                rowCursorUpdate()
            # rowCursorUpdate()
            _ = report_sheet.cell(row=rowCursorValue, column=1, value='STATE TOTAL')
            for month in month_cols:
                form = '=SUM('
                for cells in month_totals[month]:
                    form = form + cells + ','
                form = form + ')'
                _ = report_sheet.cell(row=rowCursorValue, column=month_cols.index(month) + 2, value=form)
                if month not in state_month_addresses.keys():
                    state_month_addresses[month] = []
                state_month_addresses[month].append(
                    '{}{}'.format(get_column_letter(month_cols.index(month) + 2), str(rowCursorValue)))
            _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 2,
                                  value='=SUM({}:{})'.format(
                                      'B' + str(rowCursorValue),
                                      str(get_column_letter(total_col_range + 1)) + str(
                                          rowCursorValue)))

            report_sheet.merge_cells(start_row=rowCursorValue, start_column=total_col_range + 2,
                                     end_row=rowCursorValue, end_column=total_col_range + 4)
            formula = '=SUM('
            for cell in sr_totals:
                formula = formula + cell + ','
            formula = formula + ')'
            _ = report_sheet.cell(rowCursorValue, column=total_col_range + 2, value=formula)
            state_total_addresses.append('{}{}'.format(get_column_letter(total_col_range + 2), rowCursorValue))

            for i in range(1, total_col_range + 5):
                report_sheet[get_column_letter(i) + str(rowCursorValue)].fill = state_ttl_color
            for i in range(1, rowCursorValue):
                report_sheet[get_column_letter(total_col_range + 4) + str(i)].fill = state_ttl_color
            rowCursorUpdate()
            rowCursorUpdate()
            rowCursorUpdate()
    if selection != {}:
        rowCursorUpdate()
        _ = report_sheet.cell(row=rowCursorValue, column=1, value='INDIA TOTAL')
        for i in range(1, total_col_range + 6):
            report_sheet[get_column_letter(i) + str(rowCursorValue)].fill = in_ttl_color
        for i in range(1, rowCursorValue):
            report_sheet[get_column_letter(total_col_range + 5) + str(i)].fill = in_ttl_color
        for month in month_cols:
            form = '=SUM('
            for cells in state_month_addresses[month]:
                form = form + cells + ','
            form = form + ')'
            _ = report_sheet.cell(row=rowCursorValue, column=month_cols.index(month) + 2, value=form)
        report_sheet.merge_cells(start_row=rowCursorValue, start_column=total_col_range + 2, end_row=rowCursorValue,
                                 end_column=total_col_range + 5)
        report_sheet[get_column_letter(total_col_range + 2) + str(rowCursorValue)].fill = in_color
        form = '=SUM('
        for cells in state_total_addresses:
            form = form + cells + ','
        form = form + ')'
        _ = report_sheet.cell(row=rowCursorValue, column=total_col_range + 2, value=form)

        if staged_merge_cells:
            for cell in staged_merge_cells.values():
                report_sheet.merge_cells(start_row=cell[0], start_column=cell[1], end_row=cell[2],
                                         end_column=cell[3])
        for cells in staged_merge_cells.keys():
            report_sheet[cells].alignment = Alignment(vertical='top')

    for i in range(1, total_col_range + 5):
        report_sheet[get_column_letter(i) + '1'].fill = in_color
    report_sheet.column_dimensions['A'].width = 35.78
    report_sheet.column_dimensions[get_column_letter(total_col_range + 5)].width = 10.78
    report_sheet.column_dimensions[get_column_letter(total_col_range + 4)].width = 10.78
